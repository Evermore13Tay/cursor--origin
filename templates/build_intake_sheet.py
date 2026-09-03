#!/usr/bin/env python3
"""Build the overseas intake workbook for Google Sheets import."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "overseas-intake.xlsx"
CSV_PATH = ROOT / "overseas-intake.csv"

HEADERS = [
    "submission_id",
    "submitted_at",
    "name",
    "email",
    "organization",
    "country",
    "category",
    "title",
    "details",
    "attachment_url",
    "language",
    "consent",
    "sync_status",
    "feishu_record_id",
    "last_synced_at",
    "error_message",
]

SAMPLES = [
    [
        "sub_20260903_001",
        "2026-09-03 10:00:00 UTC",
        "Alex Chen",
        "alex.chen@example.com",
        "Northwind Studio",
        "United States",
        "Partner",
        "Q4 collab brief",
        "Sharing campaign assets and launch window for the overseas test.",
        "https://drive.google.com/file/d/example",
        "en",
        "Yes",
        "New",
        "",
        "",
        "",
    ],
    [
        "sub_20260903_002",
        "2026-09-03 11:30:00 UTC",
        "Mina Sato",
        "mina.sato@example.jp",
        "Sakura Games",
        "Japan",
        "Content",
        "JP store listing copy",
        "Japanese title, subtitle, and description draft for review.",
        "https://docs.google.com/document/d/example",
        "ja",
        "Yes",
        "Synced",
        "rec_feishu_demo_002",
        "2026-09-03 12:00:00 UTC",
        "",
    ],
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
USER_FILL = PatternFill("solid", fgColor="FFF2CC")
SYNC_FILL = PatternFill("solid", fgColor="D9E2F3")
ID_FILL = PatternFill("solid", fgColor="E2EFDA")
WRAP = Alignment(vertical="center", wrap_text=True)
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
NEW_FILL = PatternFill("solid", fgColor="FCE4D6")
SYNCED_FILL = PatternFill("solid", fgColor="C6EFCE")
ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")

COL_WIDTHS = {
    "A": 20,
    "B": 24,
    "C": 18,
    "D": 28,
    "E": 22,
    "F": 16,
    "G": 14,
    "H": 28,
    "I": 48,
    "J": 42,
    "K": 12,
    "L": 12,
    "M": 16,
    "N": 22,
    "O": 22,
    "P": 28,
}

def style_header(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def paint_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row, col)
        cell.alignment = WRAP
        cell.border = THIN
        cell.font = Font(name="Calibri", size=11)
        if col == 1:
            cell.fill = ID_FILL
        elif col <= 12:
            cell.fill = USER_FILL
        else:
            cell.fill = SYNC_FILL


def add_submissions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Submissions"
    ws.sheet_properties.tabColor = "1F4E79"

    ws.append(HEADERS)
    style_header(ws, len(HEADERS))
    for sample in SAMPLES:
        ws.append(sample)
        paint_row(ws, ws.max_row, len(HEADERS))

    # Keep empty rows ready for real submissions.
    for _ in range(3, 201):
        ws.append([""] * len(HEADERS))
        paint_row(ws, ws.max_row, len(HEADERS))

    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P200"

    table = Table(displayName="SubmissionsTable", ref="A1:P200")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    category = DataValidation(
        type="list",
        formula1='"Partner,Content,Feedback,Support,Other"',
        allow_blank=True,
    )
    category.error = "Pick a category from the list."
    category.errorTitle = "Invalid category"
    category.prompt = "Partner / Content / Feedback / Support / Other"
    category.promptTitle = "Category"
    ws.add_data_validation(category)
    category.add("G2:G200")

    language = DataValidation(
        type="list",
        formula1='"en,ja,ko,es,pt,id,th,vi,other"',
        allow_blank=True,
    )
    ws.add_data_validation(language)
    language.add("K2:K200")

    consent = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(consent)
    consent.add("L2:L200")

    status = DataValidation(
        type="list",
        formula1='"New,Processing,Synced,Error,Skipped"',
        allow_blank=True,
    )
    ws.add_data_validation(status)
    status.add("M2:M200")

    ws.conditional_formatting.add(
        "M2:M200",
        FormulaRule(formula=['$M2="New"'], fill=NEW_FILL),
    )
    ws.conditional_formatting.add(
        "M2:M200",
        FormulaRule(formula=['$M2="Synced"'], fill=SYNCED_FILL),
    )
    ws.conditional_formatting.add(
        "M2:M200",
        FormulaRule(formula=['$M2="Error"'], fill=ERROR_FILL),
    )

    ws.oddHeader.center.text = "Overseas Intake  ·  sync to Feishu"
    ws.sheet_view.showGridLines = False


def add_howto(wb: Workbook) -> None:
    ws = wb.create_sheet("How to use")
    ws.sheet_properties.tabColor = "548235"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 96

    rows = [
        ("对象", "说明"),
        ("这张表是什么", "国外用户没有飞书时的中间层，对标国内腾讯在线表格。用户在 Submissions 填行，你们再同步到飞书。"),
        ("谁填黄列", "A–L 是外部用户填写区：姓名、邮箱、机构、国家、分类、标题、详情、附件链接、语言、授权。"),
        ("谁填蓝列", "M–P 是内部同步区：sync_status / feishu_record_id / last_synced_at / error_message。不要让外部用户改。"),
        ("绿列", "submission_id 是同步主键。导入飞书时用它做去重，不要用行号。"),
        ("导入 Google Sheets", "1) 打开 https://sheet.new  2) File → Import → Upload 本文件  3) 选 Replace spreadsheet。"),
        ("分享给国外用户", "Share → Anyone with the link → Editor（或 Commenter + 指定 Editor）。建议再开一个只含 A–L 的筛选视图。"),
        ("不要公开到整个互联网", "链接可编辑等于谁拿到谁都能改。正式环境请改成指定邮箱，或用 Google Form 写入本表。"),
        ("更稳的采集方式", "Insert → Form，让 Google Form 落到 Submissions。用户只需填表，不会碰同步列。"),
        ("同步到飞书", "用 Google Sheets API 或飞书企业版「数据连接」读 Submissions。按 submission_id upsert 到多维表格。"),
        ("状态约定", "New = 未同步；Processing = 同步中；Synced = 已写入飞书；Error = 失败看 error_message；Skipped = 故意不同步。"),
        ("样例行", "前两行是演示，导入后请删掉再发给外部用户。"),
    ]

    ws["A1"] = "Overseas Intake"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color="1F4E79")
    ws.merge_cells("A1:B1")
    ws["A2"] = "国外用户信息采集表  ·  腾讯表格的国际对等物"
    ws["A2"].font = Font(name="Calibri", size=12, color="595959")
    ws.merge_cells("A2:B2")

    start = 4
    ws.cell(start, 1, "对象").fill = HEADER_FILL
    ws.cell(start, 1).font = HEADER_FONT
    ws.cell(start, 2, "说明").fill = HEADER_FILL
    ws.cell(start, 2).font = HEADER_FONT
    for i, (left, right) in enumerate(rows[1:], start=start + 1):
        ws.cell(i, 1, left).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(i, 1).fill = PatternFill("solid", fgColor="DEEBF7")
        ws.cell(i, 1).alignment = Alignment(vertical="center")
        ws.cell(i, 1).border = THIN
        ws.cell(i, 2, right).font = Font(name="Calibri", size=11)
        ws.cell(i, 2).alignment = WRAP
        ws.cell(i, 2).border = THIN
        ws.row_dimensions[i].height = 36
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[4].height = 22
    ws.sheet_view.showGridLines = False


def add_field_map(wb: Workbook) -> None:
    ws = wb.create_sheet("Field map")
    ws.sheet_properties.tabColor = "C65911"
    headers = ["column", "sheet_header", "who_fills", "feishu_field_hint", "notes"]
    rows = [
        ("A", "submission_id", "User or formula", "主键 / 文本", "去重键。可用 UUID，或 email+timestamp。"),
        ("B", "submitted_at", "User or Form", "日期时间", "统一用 UTC，格式 YYYY-MM-DD HH:MM:SS UTC。"),
        ("C", "name", "User", "文本", "显示名。"),
        ("D", "email", "User", "邮箱 / 文本", "联系与去重辅助。"),
        ("E", "organization", "User", "文本", "公司 / 工作室 / 团队。"),
        ("F", "country", "User", "单选 / 文本", "国家或地区，建议英文全称。"),
        ("G", "category", "User", "单选", "Partner / Content / Feedback / Support / Other。"),
        ("H", "title", "User", "文本", "一条提交的短标题。"),
        ("I", "details", "User", "多行文本", "正文。"),
        ("J", "attachment_url", "User", "超链接 / 附件", "Google Drive / Dropbox / 图片链接。Sheets 本身不适合当网盘。"),
        ("K", "language", "User", "单选", "内容语言，方便飞书侧分流。"),
        ("L", "consent", "User", "复选框", "是否同意同步到内部系统。No 则 Skipped。"),
        ("M", "sync_status", "Internal", "单选", "同步状态机，外部只读。"),
        ("N", "feishu_record_id", "Internal", "文本", "飞书多维表格 record_id，回写用。"),
        ("O", "last_synced_at", "Internal", "日期时间", "上次成功同步时间。"),
        ("P", "error_message", "Internal", "文本", "失败原因，便于重试。"),
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    for row in rows:
        ws.append(row)
        paint_row(ws, ws.max_row, len(headers))
        ws.cell(ws.max_row, 1).fill = ID_FILL
        ws.cell(ws.max_row, 3).fill = USER_FILL if row[2].startswith("User") else SYNC_FILL
    for letter, width in {"A": 10, "B": 20, "C": 18, "D": 22, "E": 56}.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:E17"
    ws.sheet_view.showGridLines = False


def write_csv() -> None:
    import csv

    with CSV_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(SAMPLES)


def main() -> None:
    wb = Workbook()
    add_submissions(wb)
    add_howto(wb)
    add_field_map(wb)
    wb.save(XLSX_PATH)
    write_csv()
    print(f"wrote {XLSX_PATH}")
    print(f"wrote {CSV_PATH}")


if __name__ == "__main__":
    main()
